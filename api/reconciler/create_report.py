from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import logging
from api.reconciler.formatting import apply_cell_formatting


# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
BOLD_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True, color="000000")
NUMBER_FORMAT = "#,##0.00"
REMARKS_COLUMN = 5  # Assuming the remarks column is the 5th column (E
# Cell formatting
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

COLORS = {
    'MATCHED': PatternFill(start_color="5C7AFF", end_color="5C7AFF", fill_type="solid"),
    'FUZZY': PatternFill(start_color="B39CD0", end_color="B39CD0", fill_type="solid"),
    'SPLIT': PatternFill(start_color="59D2FE", end_color="59D2FE", fill_type="solid"),
    'RETURNED': PatternFill(start_color="44E5E7", end_color="44E5E7", fill_type="solid"),
    'ROUNDING': PatternFill(start_color="73FBD3", end_color="73FBD3", fill_type="solid"),
    'UNMATCHED': PatternFill(start_color="8EC1FF", end_color="8EC1FF", fill_type="solid"),
    'CLOSING_MATCHED': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'CLOSING_UNMATCHED': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'TOTAL_MATCHED': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'TOTAL_UNMATCHED': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'HEADER': PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
}


def create_reconciliation_report(wb,
                                 matched_rows1, matched_rows2, fuzzy_rows1, fuzzy_rows2,
                                 split_rows1, split_rows2, returned_rows1, returned_rows2,
                                 rounding_rows1, rounding_rows2, unmatched_rows1, unmatched_rows2,
                                 closing_debit1, closing_credit1, closing_debit2, closing_credit2,
                                 closing_match, total_match):
    """
    Create a professionally formatted reconciliation report worksheet.
    """
    if "Reconciliation Report" in wb.sheetnames:
        ws_report = wb["Reconciliation Report"]
    else:
        ws_report = wb.create_sheet("Reconciliation Report")
    
    # Clear existing content
    for row in ws_report.rows:
        for cell in row:
            cell.value = None

    # Set column widths
    ws_report.column_dimensions['A'].width = 30
    ws_report.column_dimensions['B'].width = 22
    ws_report.column_dimensions['C'].width = 22

    # Merge cells for main header
    ws_report.merge_cells('A1:C1')
    header_cell = ws_report.cell(row=1, column=1, value="RECONCILIATION REPORT")
    header_cell.font = Font(bold=True, size=14, color="000000")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_cell.border = Border(
        left=Side(style="medium"), right=Side(style="medium"),
        top=Side(style="medium"), bottom=Side(style="medium")
    )
    
    # Empty row for spacing
    ws_report.append(["", "", ""])
    
    # Category headers
    category_row = 3
    ws_report.cell(row=category_row, column=1, value="Category").font = BOLD_FONT
    ws_report.cell(row=category_row, column=2, value="Sheet1 Count").font = BOLD_FONT
    ws_report.cell(row=category_row, column=3, value="Sheet2 Count").font = BOLD_FONT
    for col in range(1, 4):
        cell = ws_report.cell(row=category_row, column=col)
        cell.fill = COLORS['HEADER']
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        cell.alignment = CENTER_ALIGN
    
    # Match statistics
    data = [
        ["Matched (Exact)", len(matched_rows1), len(matched_rows2)],
        ["Matched but check date", len(fuzzy_rows1), len(fuzzy_rows2)],
        ["Split Transaction", len(split_rows1), len(split_rows2)],
        ["Returned Transaction", len(returned_rows1)//2, len(returned_rows2)//2],
        ["Rounding Error", len(rounding_rows1), len(rounding_rows2)],
        ["Unmatched", len(unmatched_rows1), len(unmatched_rows2)]
    ]
    
    for idx, row_data in enumerate(data, 4):
        for col, value in enumerate(row_data, 1):
            cell = ws_report.cell(row=idx, column=col, value=value)
            cell.border = THIN_BORDER
            if col == 1:
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = CENTER_ALIGN
                cell.number_format = "0"
    
    # Space before Balance Information
    row = 10
    ws_report.merge_cells(f'A{row}:C{row}')
    
    # Balance Information header
    row = 11
    balance_header_cell = ws_report.cell(row=row, column=1, value="BALANCE INFORMATION")
    balance_header_cell.font = Font(bold=True, size=12, color="000000")
    balance_header_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    balance_header_cell.alignment = LEFT_ALIGN
    ws_report.merge_cells(f'A{row}:C{row}')
    
    # Closing balance status
    row = 13
    match_status = "MATCHED" if closing_match else "UNMATCHED"
    status_fill = COLORS['CLOSING_MATCHED'] if closing_match else COLORS['CLOSING_UNMATCHED']
    status_label = ws_report.cell(row=row, column=1, value="Closing Balance Status:")
    status_label.font = BOLD_FONT
    status_label.border = THIN_BORDER
    status_label.alignment = LEFT_ALIGN
    status_cell = ws_report.cell(row=row, column=2, value=match_status)
    status_cell.fill = status_fill
    status_cell.border = THIN_BORDER
    status_cell.alignment = CENTER_ALIGN
    status_cell.font = BOLD_FONT
    empty_cell = ws_report.cell(row=row, column=3)
    empty_cell.border = THIN_BORDER
    
    # Total status
    row = 14
    total_status = "MATCHED" if total_match else "UNMATCHED"
    total_fill = COLORS['TOTAL_MATCHED'] if total_status == "MATCHED" else COLORS['TOTAL_UNMATCHED']
    total_label = ws_report.cell(row=row, column=1, value="Total Status:")
    total_label.font = BOLD_FONT
    total_label.border = THIN_BORDER
    total_label.alignment = LEFT_ALIGN
    total_cell = ws_report.cell(row=row, column=2, value=total_status)
    total_cell.fill = total_fill
    total_cell.border = THIN_BORDER
    total_cell.alignment = CENTER_ALIGN
    total_cell.font = BOLD_FONT
    empty_cell = ws_report.cell(row=row, column=3)
    empty_cell.border = THIN_BORDER
    
    # Space before Color Legend
    row = 16
    ws_report.merge_cells(f'A{row}:C{row}')
    
    # Legend header
    row = 17
    legend_header_cell = ws_report.cell(row=row, column=1, value="COLOR LEGEND")
    legend_header_cell.font = Font(bold=True, size=12, color="000000")
    legend_header_cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    legend_header_cell.alignment = LEFT_ALIGN
    ws_report.merge_cells(f'A{row}:C{row}')
    
    # Legend items
    legend_items = [
        ("Matched", COLORS['MATCHED']),
        ("Matched but check date", COLORS['FUZZY']),
        ("Split Transaction", COLORS['SPLIT']),
        ("Returned Transaction", COLORS['RETURNED']),
        ("Rounding Error", COLORS['ROUNDING']),
        ("Unmatched", COLORS['UNMATCHED']),
        ("Closing Balance Matched", COLORS['CLOSING_MATCHED']),
        ("Closing Balance Unmatched", COLORS['CLOSING_UNMATCHED'])
    ]
    
    legend_start_row = 19
    for i, (label, fill) in enumerate(legend_items):
        row = legend_start_row + i
        label_cell = ws_report.cell(row=row, column=1, value=label)
        label_cell.border = THIN_BORDER
        label_cell.alignment = LEFT_ALIGN
        color_cell = ws_report.cell(row=row, column=2)
        color_cell.fill = fill
        color_cell.border = THIN_BORDER
        empty_cell = ws_report.cell(row=row, column=3)
        empty_cell.border = THIN_BORDER
    
    # Alternate row shading for stats & legend
    for row in range(4, 10):
        if row % 2 == 0:
            for col in range(1, 4):
                cell = ws_report.cell(row=row, column=col)
                if not cell.fill.start_color.index or cell.fill.start_color.index == "00000000":
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for row in range(legend_start_row, legend_start_row + len(legend_items)):
        if row % 2 == 0:
            for col in [1, 3]:
                cell = ws_report.cell(row=row, column=col)
                if not cell.fill.start_color.index or cell.fill.start_color.index == "00000000":
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    logger.info("Created professional reconciliation report.")

def add_closing_and_total_rows(ws, last_data_row, data_start_row, closing_match, total_match):
    closing_row = last_data_row + 1
    total_row = last_data_row + 2

    # Closing Balance
    ws.cell(row=closing_row, column=2, value="Closing Balance").font = BOLD_FONT
    ws.cell(row=closing_row, column=3).value = (
        f'=IF(SUM(D{data_start_row}:D{last_data_row})>SUM(C{data_start_row}:C{last_data_row}),'
        f'SUM(D{data_start_row}:D{last_data_row})-SUM(C{data_start_row}:C{last_data_row}),0)'
    )
    ws.cell(row=closing_row, column=4).value = (
        f'=IF(SUM(C{data_start_row}:C{last_data_row})>SUM(D{data_start_row}:D{last_data_row}),'
        f'SUM(C{data_start_row}:C{last_data_row})-SUM(D{data_start_row}:D{last_data_row}),0)'
    )
    
    status_text = "Closing Balance: Matched" if closing_match else "Closing Balance: Unmatched"
    ws.cell(row=closing_row, column=REMARKS_COLUMN, value=status_text).font = BOLD_FONT
    status_fill = COLORS['CLOSING_MATCHED'] if closing_match else COLORS['CLOSING_UNMATCHED']
    apply_cell_formatting(ws, closing_row, range(1, REMARKS_COLUMN + 1),
                          fill=status_fill, border=THIN_BORDER, number_format=NUMBER_FORMAT)
    
    # Total
    ws.cell(row=total_row, column=2, value="Total").font = BOLD_FONT
    ws.cell(row=total_row, column=3).value = f"=SUM(C{data_start_row}:C{closing_row})"
    ws.cell(row=total_row, column=4).value = f"=SUM(D{data_start_row}:D{closing_row})"
    
    status_text = "Total: Matched" if total_match else "Total: Unmatched"
    ws.cell(row=total_row, column=REMARKS_COLUMN, value=status_text).font = BOLD_FONT
    status_fill = COLORS['TOTAL_MATCHED'] if total_match else COLORS['TOTAL_UNMATCHED']
    apply_cell_formatting(ws, total_row, range(1, REMARKS_COLUMN + 1),
                          fill=status_fill, border=THIN_BORDER, number_format=NUMBER_FORMAT)

    return closing_row, total_row

def apply_professional_formatting(ws1, ws2, header_row, data_start_row, last_row1, last_row2):
    # Column widths
    for ws in [ws1, ws2]:
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30

    # Header row formatting
    for ws in [ws1, ws2]:
        for col in range(1, REMARKS_COLUMN + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = HEADER_FONT
            cell.fill = COLORS['HEADER']
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

    # Data rows formatting
    for ws, last_row in [(ws1, last_row1), (ws2, last_row2)]:
        for row in range(data_start_row, last_row + 1):
            apply_cell_formatting(ws, row, [1], fill=None, align=CENTER_ALIGN, border=THIN_BORDER)
            apply_cell_formatting(ws, row, [2], fill=None, align=LEFT_ALIGN, border=THIN_BORDER)
            apply_cell_formatting(ws, row, [3, 4], fill=None, align=RIGHT_ALIGN,
                                  border=THIN_BORDER, number_format=NUMBER_FORMAT)
            apply_cell_formatting(ws, row, [REMARKS_COLUMN], fill=None, align=LEFT_ALIGN,
                                  border=THIN_BORDER)

    # Company/Ledger name cells
    for ws in [ws1, ws2]:
        for row in range(1, 3):
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                if col == 1:
                    cell.font = BOLD_FONT
                cell.border = THIN_BORDER
                cell.alignment = LEFT_ALIGN