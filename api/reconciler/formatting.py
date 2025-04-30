import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

REMARKS_COLUMN = 5  # Assuming the remarks column is the 5th column (E)

# Color definitions
COLORS = {
    'MATCHED': PatternFill(start_color="5C7AFF", end_color="5C7AFF", fill_type="solid"),
    'FUZZY': PatternFill(start_color="B39CD0", end_color="B39CD0", fill_type="solid"),
    'SPLIT': PatternFill(start_color="59D2FE", end_color="59D2FE", fill_type="solid"),
    'RETURNED': PatternFill(start_color="44E5E7", end_color="44E5E7", fill_type="solid"),
    'ROUNDING': PatternFill(start_color="73FBD3", end_color="73FBD3", fill_type="solid"),
    'UNMATCHED': PatternFill(start_color="8EC1FF", end_color="8EC1FF", fill_type="solid"),
    'CLOSING_MATCHED': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'CLOSING_UNMATCHED': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'TOTAL_MATCHED': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'TOTAL_UNMATCHED': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'HEADER': PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
}


def apply_cell_formatting(ws, row, col_range, fill, font=None, align=None, number_format=None, border=None):
    for col in col_range:
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        if number_format:
            cell.number_format = number_format
        if border:
            cell.border = border

def write_remarks_to_sheets(df1, df2, data_start_row, ws1, ws2):
    unmatched_rows1 = []
    unmatched_rows2 = []
    for i in df1.index:
        ws1.cell(row=i + data_start_row, column=REMARKS_COLUMN, value=df1.at[i, "Remarks"])
        if df1.at[i, "Remarks"] == "Unmatched":
            unmatched_rows1.append(i + data_start_row)
    for j in df2.index:
        ws2.cell(row=j + data_start_row, column=REMARKS_COLUMN, value=df2.at[j, "Remarks"])
        if df2.at[j, "Remarks"] == "Unmatched":
            unmatched_rows2.append(j + data_start_row)
    return unmatched_rows1, unmatched_rows2

def apply_color_formatting(ws1, ws2, matched_rows1, matched_rows2, fuzzy_rows1, fuzzy_rows2,
                           split_rows1, split_rows2, returned_rows1, returned_rows2,
                           rounding_rows1, rounding_rows2, unmatched_rows1, unmatched_rows2):
    def color_row(ws, row, fill):
        for col in range(1, REMARKS_COLUMN + 1):
            ws.cell(row=row, column=col).fill = fill

    # Sheet1
    for row in matched_rows1:
        color_row(ws1, row, COLORS['MATCHED'])
    for row in fuzzy_rows1:
        color_row(ws1, row, COLORS['FUZZY'])
    for row in split_rows1:
        color_row(ws1, row, COLORS['SPLIT'])
    for row in returned_rows1:
        color_row(ws1, row, COLORS['RETURNED'])
    for row in rounding_rows1:
        color_row(ws1, row, COLORS['ROUNDING'])
    for row in unmatched_rows1:
        color_row(ws1, row, COLORS['UNMATCHED'])

    # Sheet2
    for row in matched_rows2:
        color_row(ws2, row, COLORS['MATCHED'])
    for row in fuzzy_rows2:
        color_row(ws2, row, COLORS['FUZZY'])
    for row in split_rows2:
        color_row(ws2, row, COLORS['SPLIT'])
    for row in returned_rows2:
        color_row(ws2, row, COLORS['RETURNED'])
    for row in rounding_rows2:
        color_row(ws2, row, COLORS['ROUNDING'])
    for row in unmatched_rows2:
        color_row(ws2, row, COLORS['UNMATCHED'])