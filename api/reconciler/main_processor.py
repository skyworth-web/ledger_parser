#!/usr/bin/env python
import os
import math
import logging
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Font, Border, Protection, NamedStyle
from api.reconciler.matchers import (find_exact_matches, find_fuzzy_matches,
                      find_split_transactions, find_rounding_errors)
from api.reconciler.utils import compare_values, calculate_closing_balance
from api.reconciler.formatting import (
    apply_cell_formatting, write_remarks_to_sheets, apply_color_formatting)
from api.reconciler.create_report import create_reconciliation_report, add_closing_and_total_rows, apply_professional_formatting
from api.reconciler.config_utils import load_config
from copy import copy  # Add this import at the top
config = load_config()

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Sheet structure
HEADER_ROW = 4
HEADER_ROW = 6  # 1-based index
DATA_START_ROW = 7
EXPECTED_COLUMNS = ["date", "description", "debit", "credit"]
REMARKS_COLUMN = 5  # Column E

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
BOLD_FONT = Font(bold=True)


def copy_worksheet(source_ws, target_wb, title):
    target_ws = target_wb.create_sheet(title=title)
    
    for row in source_ws.iter_rows():
        for cell in row:
            # Corrected 'cell.col_idx' to 'cell.column'
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            
            # Copy cell styles
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy(cell.alignment)
    
    return target_ws


def find_returned_transactions(df, unmatched_set, data_start_row, returned_rows):
    returned_count = 0
    indices = sorted(list(unmatched_set))
    marked = set()
    for idx_i in indices:
        if idx_i in marked:
            continue
        for idx_j in indices:
            if idx_j <= idx_i or idx_j in marked:
                continue
            date_i = df.at[idx_i, "date"]
            date_j = df.at[idx_j, "date"]
            if isinstance(date_i, pd.Timestamp) and isinstance(date_j, pd.Timestamp):
                date_diff = abs((date_i - date_j).days)
                date_match = date_diff <= 1
            else:
                date_match = (date_i == date_j)
            if date_match:
                # reversed transactions
                if (df.at[idx_i, "debit"] > 0 and df.at[idx_i, "credit"] == 0 and
                    df.at[idx_j, "credit"] > 0 and df.at[idx_j, "debit"] == 0 and
                    compare_values(df.at[idx_i, "debit"], df.at[idx_j, "credit"])):
                    df.at[idx_i, "Remarks"] = "Returned Transaction"
                    df.at[idx_j, "Remarks"] = "Returned Transaction"
                    returned_rows.append(idx_i + data_start_row)
                    returned_rows.append(idx_j + data_start_row)
                    marked.add(idx_i)
                    marked.add(idx_j)
                    returned_count += 1
                    break
                elif (df.at[idx_i, "credit"] > 0 and df.at[idx_i, "debit"] == 0 and
                      df.at[idx_j, "debit"] > 0 and df.at[idx_j, "credit"] == 0 and
                      compare_values(df.at[idx_i, "credit"], df.at[idx_j, "debit"])):
                    df.at[idx_i, "Remarks"] = "Returned Transaction"
                    df.at[idx_j, "Remarks"] = "Returned Transaction"
                    returned_rows.append(idx_i + data_start_row)
                    returned_rows.append(idx_j + data_start_row)
                    marked.add(idx_i)
                    marked.add(idx_j)
                    returned_count += 1
                    break
    unmatched_set.difference_update(marked)
    logger.info(f"Found {returned_count} returned transactions.")
    return returned_count


def reconcile_statement(file_path1, file_path2):
    try:
        wb1 = load_workbook(file_path1)
        wb2 = load_workbook(file_path2)
        source_ws1 = wb1["Sheet1"]
        source_ws2 = wb2["Sheet1"]
    except Exception as e:
        logger.error(f"Error loading input files: {e}")
        return False

    try:
        df1 = pd.read_excel(file_path1, sheet_name="Sheet1", header=HEADER_ROW - 1)
        df2 = pd.read_excel(file_path2, sheet_name="Sheet1", header=HEADER_ROW - 1)
    except Exception as e:
        logger.error(f"Error reading sheets into DataFrames: {e}")
        return False

    missing_cols_df1 = [col for col in EXPECTED_COLUMNS if col not in df1.columns]
    missing_cols_df2 = [col for col in EXPECTED_COLUMNS if col not in df2.columns]
    if missing_cols_df1:
        logger.error(f"Ledger1 is missing columns: {missing_cols_df1}.")
        return False
    if missing_cols_df2:
        logger.error(f"Ledger2 is missing columns: {missing_cols_df2}.")
        return False

    df1 = df1[EXPECTED_COLUMNS].copy()
    df2 = df2[EXPECTED_COLUMNS].copy()

    df1["date"] = pd.to_datetime(df1["date"], errors="coerce")
    df2["date"] = pd.to_datetime(df2["date"], errors="coerce")
    df1["debit"] = df1["debit"].fillna(0)
    df1["credit"] = df1["credit"].fillna(0)
    df2["debit"] = df2["debit"].fillna(0)
    df2["credit"] = df2["credit"].fillna(0)

    df1["Remarks"] = ""
    df2["Remarks"] = ""

    wb = Workbook()
    wb.remove(wb.active)
    ws1 = copy_worksheet(source_ws1, wb, "Sheet1")
    ws2 = copy_worksheet(source_ws2, wb, "Sheet2")

    ws1.cell(row=HEADER_ROW, column=REMARKS_COLUMN, value="Remarks").font = BOLD_FONT
    ws2.cell(row=HEADER_ROW, column=REMARKS_COLUMN, value="Remarks").font = BOLD_FONT

    unmatched_df1 = set(df1.index)
    unmatched_df2 = set(df2.index)
    matched_rows1 = []
    matched_rows2 = []
    fuzzy_rows1 = []
    fuzzy_rows2 = []
    split_rows1 = []
    split_rows2 = []
    returned_rows1 = []
    returned_rows2 = []
    rounding_rows1 = []
    rounding_rows2 = []

    find_exact_matches(df1, df2, unmatched_df1, unmatched_df2, DATA_START_ROW,
                       matched_rows1, matched_rows2, fuzzy_rows1, fuzzy_rows2, config)

    find_fuzzy_matches(df1, df2, unmatched_df1, unmatched_df2, DATA_START_ROW,
                       matched_rows1, matched_rows2, fuzzy_rows1, fuzzy_rows2, config)

    find_split_transactions(df1, df2, unmatched_df1, unmatched_df2, DATA_START_ROW,
                            split_rows1, split_rows2, config)
    find_split_transactions(df2, df1, unmatched_df2, unmatched_df1, DATA_START_ROW,
                            split_rows2, split_rows1, config)

    find_rounding_errors(df1, df2, unmatched_df1, unmatched_df2, DATA_START_ROW,
                         rounding_rows1, rounding_rows2, config)

    find_returned_transactions(df1, unmatched_df1, DATA_START_ROW, returned_rows1)
    find_returned_transactions(df2, unmatched_df2, DATA_START_ROW, returned_rows2)

    for i in unmatched_df1:
        df1.at[i, "Remarks"] = "Unmatched"
    for j in unmatched_df2:
        df2.at[j, "Remarks"] = "Unmatched"

    unmatched_rows1, unmatched_rows2 = write_remarks_to_sheets(df1, df2, DATA_START_ROW, ws1, ws2)

    apply_color_formatting(ws1, ws2,
                           matched_rows1, matched_rows2,
                           fuzzy_rows1, fuzzy_rows2,
                           split_rows1, split_rows2,
                           returned_rows1, returned_rows2,
                           rounding_rows1, rounding_rows2,
                           unmatched_rows1, unmatched_rows2)

    closing_debit1, closing_credit1 = calculate_closing_balance(df1)
    closing_debit2, closing_credit2 = calculate_closing_balance(df2)

    closing_match = (compare_values(closing_debit1, closing_credit2) and
                     compare_values(closing_credit1, closing_debit2))

    total_debit1 = df1["debit"].sum() + closing_debit1
    total_credit1 = df1["credit"].sum() + closing_credit1
    total_debit2 = df2["debit"].sum() + closing_debit2
    total_credit2 = df2["credit"].sum() + closing_credit2
    total_match = (compare_values(total_debit1, total_debit2) and
                   compare_values(total_credit1, total_credit2))

    last_data_row1 = DATA_START_ROW + len(df1) - 1
    last_data_row2 = DATA_START_ROW + len(df2) - 1

    closing_row1, total_row1 = add_closing_and_total_rows(ws1, last_data_row1, DATA_START_ROW,
                                                          closing_match, total_match)
    closing_row2, total_row2 = add_closing_and_total_rows(ws2, last_data_row2, DATA_START_ROW,
                                                          closing_match, total_match)

    apply_professional_formatting(ws1, ws2, HEADER_ROW, DATA_START_ROW, total_row1, total_row2)

    ws1.auto_filter.ref = f"A{HEADER_ROW}:E{total_row1}"
    ws2.auto_filter.ref = f"A{HEADER_ROW}:E{total_row2}"

    create_reconciliation_report(
        wb,
        matched_rows1, matched_rows2,
        fuzzy_rows1, fuzzy_rows2,
        split_rows1, split_rows2,
        returned_rows1, returned_rows2,
        rounding_rows1, rounding_rows2,
        unmatched_rows1, unmatched_rows2,
        closing_debit1, closing_credit1,
        closing_debit2, closing_credit2,
        closing_match, total_match
    )

    try:
        current_time = datetime.now()
        time_str = current_time.strftime("_%Y%m%d_%H%M%S")
        output_path = './data/output/reconciled/' + "reconciled" + time_str + '.xlsx'
        wb.save(output_path)     
        return output_path
    except Exception as e:
        logger.error(f"Error saving workbook: {e}")
        return False
